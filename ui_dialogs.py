import os
import time
import subprocess
import logging

import win32api
import win32print

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton, QDialogButtonBox,
    QMessageBox, QTabWidget, QHBoxLayout, QTableWidget, QTableWidgetItem
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ui_delegates import RotatedHeaderDelegate


class PrinterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wybierz drukarkę")
        self.resize(400, 200)
        self.selected_printer = None
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Wybierz urządzenie do wydruku:"))
        self.printer_combo = QComboBox()
        self.printers = self.get_system_printers()
        default_printer = win32print.GetDefaultPrinter()
        for p in self.printers:
            self.printer_combo.addItem(p)
            if p == default_printer:
                self.printer_combo.setCurrentText(p)
        layout.addWidget(self.printer_combo)
        self.btn_properties = QPushButton("Sprawdź ustawienia / Włącz Duplex")
        self.btn_properties.clicked.connect(self.open_printer_properties)
        layout.addWidget(self.btn_properties)
        layout.addWidget(
            QLabel("<i>Wskazówka: Kliknij powyżej, aby upewnić się,<br>że druk dwustronny jest włączony.</i>"))
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_system_printers(self):
        printers = []
        try:
            flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
            for p in win32print.EnumPrinters(flags):
                printers.append(p[2])
        except Exception:
            pass
        return printers

    def open_printer_properties(self):
        printer_name = self.printer_combo.currentText()
        if not printer_name:
            return
        try:
            cmd = f'rundll32 printui.dll,PrintUIEntry /p /n "{printer_name}"'
            subprocess.Popen(cmd, shell=True)
        except Exception as e:
            QMessageBox.warning(self, "Błąd", f"Nie udało się otworzyć ustawień: {e}")

    def accept(self):
        self.selected_printer = self.printer_combo.currentText()
        super().accept()


class ExcelPreviewDialog(QDialog):
    def __init__(self, filepath, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Podgląd: {os.path.basename(filepath)}")
        self.resize(1100, 800)
        self.filepath = filepath

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Podgląd (Przełączaj zakładki poniżej):"))

        # Zakładki dla arkuszy
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # --- ZMIANA: Dodano przycisk drukowania w oknie podglądu ---
        btn_layout = QHBoxLayout()

        self.print_btn = QPushButton("🖨️ Drukuj ten dokument")
        self.print_btn.setMinimumHeight(40)
        self.print_btn.setStyleSheet("font-weight: bold; background-color: #2196F3; color: white;")
        self.print_btn.clicked.connect(self.print_current_file)

        close_btn = QPushButton("Zamknij")
        close_btn.setMinimumHeight(40)
        close_btn.clicked.connect(self.reject)

        btn_layout.addWidget(self.print_btn)
        btn_layout.addStretch()  # Odstęp
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

        self.setLayout(layout)

        self.rotated_delegate = RotatedHeaderDelegate(self)
        self.load_excel_data()

    def load_excel_data(self):
        try:
            wb = load_workbook(self.filepath, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = QTableWidget()
                self.sheet_to_table(ws, table)
                self.tabs.addTab(table, sheet_name)

            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "Błąd podglądu", f"Nie udało się wczytać pliku:\n{e}")

    def sheet_to_table(self, ws, table):
        max_r = ws.max_row
        max_c = ws.max_column

        # Wymuszenie 18 kolumn dla Ewidencji
        if "Ewidencja" in ws.title and max_c < 18:
            max_c = 18

        table.setRowCount(max_r)
        table.setColumnCount(max_c)
        table.horizontalHeader().setVisible(False)
        table.verticalHeader().setVisible(False)
        table.setStyleSheet("QTableWidget { background-color: white; gridline-color: #a0a0a0; }")

        # Wymiary
        for col_idx in range(1, max_c + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                if width:
                    table.setColumnWidth(col_idx - 1, int(width * 7.5))

        for row_idx in range(1, max_r + 1):
            if row_idx in ws.row_dimensions:
                height = ws.row_dimensions[row_idx].height
                if height:
                    table.setRowHeight(row_idx - 1, int(height * 1.33))

        # Iteracja
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                str_val = str(val) if val is not None else ""

                item = QTableWidgetItem(str_val)
                item.setFlags(Qt.ItemIsEnabled)

                # Styl
                font = QFont("Times New Roman", 9)
                if cell.font:
                    if cell.font.name:
                        font.setFamily(cell.font.name)
                    if cell.font.sz:
                        font.setPointSize(int(cell.font.sz))
                    if cell.font.b:
                        font.setBold(True)
                item.setFont(font)

                align = Qt.AlignVCenter
                if cell.alignment:
                    if cell.alignment.horizontal == 'center':
                        align |= Qt.AlignHCenter
                    elif cell.alignment.horizontal == 'right':
                        align |= Qt.AlignRight
                    else:
                        align |= Qt.AlignLeft
                    if cell.alignment.vertical == 'top':
                        align = (align & ~Qt.AlignVCenter) | Qt.AlignTop
                item.setTextAlignment(align)

                # Tło
                if cell.fill and cell.fill.patternType == 'solid':
                    fg = cell.fill.fgColor
                    if hasattr(fg, 'rgb') and fg.rgb:
                        if isinstance(fg.rgb, str):
                            hex_color = "#" + fg.rgb[2:] if len(fg.rgb) > 6 else "#" + fg.rgb
                            item.setBackground(QColor(hex_color))

                table.setItem(r - 1, c - 1, item)

                # ROTACJA
                if cell.alignment and cell.alignment.textRotation:
                    if cell.alignment.textRotation in [90, 180]:
                        table.setItemDelegateForColumn(c - 1, self.rotated_delegate)

        # Scalanie
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            table.setSpan(min_row - 1, min_col - 1,
                          max_row - min_row + 1, max_col - min_col + 1)

    # --- NOWA METODA DRUKOWANIA Z PODGLĄDU ---
    def print_current_file(self):
        printer_dialog = PrinterDialog(self)
        if printer_dialog.exec_() == QDialog.Accepted:
            selected_printer = printer_dialog.selected_printer
            original_printer = win32print.GetDefaultPrinter()
            try:
                win32print.SetDefaultPrinter(selected_printer)
                # Drukujemy aktualnie podglądany plik
                win32api.ShellExecute(0, "print", self.filepath, None, ".", 0)
                QMessageBox.information(self, "Sukces", f"Wysłano do druku:\n{selected_printer}")
            except Exception as e:
                logging.error(f"Błąd druku z podglądu: {e}")
                QMessageBox.critical(self, "Błąd druku", f"Wystąpił błąd: {e}")
            finally:
                win32print.SetDefaultPrinter(original_printer)


class EdytorPracownikow(QDialog):
    def __init__(self, current_text, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edytuj listę pracowników")
        self.resize(400, 300)
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Wpisz pracowników (Imię Nazwisko, Stanowisko):"))
        self.text_edit = __import__("PyQt5.QtWidgets").QtWidgets.QTextEdit()
        self.text_edit.setPlainText(current_text)
        layout.addWidget(self.text_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText()
