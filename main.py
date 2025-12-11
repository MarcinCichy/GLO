import sys
import os
import time
import subprocess
import holidays
# Import konieczny dla pliku EXE
import holidays.countries.poland
import calendar
import win32api
import win32print
import configparser
import logging
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QComboBox, QTextEdit,
                             QPushButton, QGroupBox, QCalendarWidget, QTableView,
                             QMessageBox, QFileDialog, QListWidget, QListWidgetItem,
                             QDialog, QDialogButtonBox, QAbstractItemView, QProgressBar,
                             QTableWidget, QTableWidgetItem, QTabWidget, QHeaderView,
                             QStyledItemDelegate, QStyle)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal, QRectF
from PyQt5.QtGui import QColor, QPainter, QIcon, QFont, QBrush, QPalette

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# === KONFIGURACJA LOGOWANIA B≈ÅƒòD√ìW ===
logging.basicConfig(filename='debug.log', level=logging.ERROR,
                    format='%(asctime)s:%(levelname)s:%(message)s')


# ============================================================================
# 1. KLASY POMOCNICZE (W G√ìRNEJ CZƒò≈öCI PLIKU - ≈ªEBY UNIKNƒÑƒÜ B≈ÅƒòD√ìW)
# ============================================================================

class RotatedHeaderDelegate(QStyledItemDelegate):
    """
    Rysuje pionowy tekst w tabeli podglƒÖdu.
    Naprawiono b≈ÇƒÖd: Nie obraca tekstu w stopce (wiersze > 33).
    """

    def paint(self, painter, option, index):
        # Je≈õli to wiersz stopki (indeks > 32, czyli wiersz 34+), rysuj normalnie
        if index.row() > 32:
            QStyledItemDelegate.paint(self, painter, option, index)
            return

        text = index.data(Qt.DisplayRole)
        bg_brush = index.data(Qt.BackgroundRole)

        painter.save()

        # 1. T≈Ço
        if bg_brush:
            painter.fillRect(option.rect, bg_brush)

        # 2. Ramka
        painter.setPen(QColor("#dcdcdc"))
        painter.drawRect(option.rect)

        # 3. Tekst
        if text:
            painter.setFont(option.font)
            painter.setPen(Qt.black)

            # Translacja ≈õrodka i rotacja -90
            rect = option.rect
            painter.translate(rect.center())
            painter.rotate(-90)

            # Po obrocie zamieniamy wymiary (W, H -> H, W)
            text_rect = QRectF(-rect.height() / 2, -rect.width() / 2, rect.height(), rect.width())
            painter.drawText(text_rect, Qt.AlignCenter | Qt.TextWordWrap, text)

        painter.restore()


class PrinterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wybierz drukarkƒô")
        self.resize(400, 200)
        self.selected_printer = None
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Wybierz urzƒÖdzenie do wydruku:"))
        self.printer_combo = QComboBox()
        self.printers = self.get_system_printers()
        default_printer = win32print.GetDefaultPrinter()
        for p in self.printers:
            self.printer_combo.addItem(p)
            if p == default_printer:
                self.printer_combo.setCurrentText(p)
        layout.addWidget(self.printer_combo)
        self.btn_properties = QPushButton("Sprawd≈∫ ustawienia / W≈ÇƒÖcz Duplex")
        self.btn_properties.clicked.connect(self.open_printer_properties)
        layout.addWidget(self.btn_properties)
        layout.addWidget(
            QLabel("<i>Wskaz√≥wka: Kliknij powy≈ºej, aby upewniƒá siƒô,<br>≈ºe druk dwustronny jest w≈ÇƒÖczony.</i>"))
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_system_printers(self):
        printers = []
        try:
            flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
            for p in win32print.EnumPrinters(flags):
                printers.append(p[2])
        except Exception:
            pass
        return printers

    def open_printer_properties(self):
        printer_name = self.printer_combo.currentText()
        if not printer_name: return
        try:
            cmd = f'rundll32 printui.dll,PrintUIEntry /p /n "{printer_name}"'
            subprocess.Popen(cmd, shell=True)
        except Exception as e:
            QMessageBox.warning(self, "B≈ÇƒÖd", f"Nie uda≈Ço siƒô otworzyƒá ustawie≈Ñ: {e}")

    def accept(self):
        self.selected_printer = self.printer_combo.currentText()
        super().accept()


class ExcelPreviewDialog(QDialog):
    def __init__(self, filepath, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"PodglƒÖd: {os.path.basename(filepath)}")
        self.resize(1100, 800)
        self.filepath = filepath

        layout = QVBoxLayout()
        layout.addWidget(QLabel("PodglƒÖd (Prze≈ÇƒÖczaj zak≈Çadki poni≈ºej):"))

        # Zak≈Çadki dla arkuszy
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # --- ZMIANA: Dodano przycisk drukowania w oknie podglƒÖdu ---
        btn_layout = QHBoxLayout()

        self.print_btn = QPushButton("üñ®Ô∏è Drukuj ten dokument")
        self.print_btn.setMinimumHeight(40)
        self.print_btn.setStyleSheet("font-weight: bold; background-color: #2196F3; color: white;")
        self.print_btn.clicked.connect(self.print_current_file)

        close_btn = QPushButton("Zamknij")
        close_btn.setMinimumHeight(40)
        close_btn.clicked.connect(self.reject)

        btn_layout.addWidget(self.print_btn)
        btn_layout.addStretch()  # Odstƒôp
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

        self.setLayout(layout)

        self.rotated_delegate = RotatedHeaderDelegate(self)
        self.load_excel_data()

    def load_excel_data(self):
        try:
            wb = load_workbook(self.filepath, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = QTableWidget()
                self.sheet_to_table(ws, table)
                self.tabs.addTab(table, sheet_name)

            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "B≈ÇƒÖd podglƒÖdu", f"Nie uda≈Ço siƒô wczytaƒá pliku:\n{e}")

    def sheet_to_table(self, ws, table):
        max_r = ws.max_row
        max_c = ws.max_column

        # Wymuszenie 18 kolumn dla Ewidencji
        if "Ewidencja" in ws.title and max_c < 18:
            max_c = 18

        table.setRowCount(max_r)
        table.setColumnCount(max_c)
        table.horizontalHeader().setVisible(False)
        table.verticalHeader().setVisible(False)
        table.setStyleSheet("QTableWidget { background-color: white; gridline-color: #a0a0a0; }")

        # Wymiary
        for col_idx in range(1, max_c + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                if width:
                    table.setColumnWidth(col_idx - 1, int(width * 7.5))

        for row_idx in range(1, max_r + 1):
            if row_idx in ws.row_dimensions:
                height = ws.row_dimensions[row_idx].height
                if height:
                    table.setRowHeight(row_idx - 1, int(height * 1.33))

        # Iteracja
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                str_val = str(val) if val is not None else ""

                item = QTableWidgetItem(str_val)
                item.setFlags(Qt.ItemIsEnabled)

                # Styl
                font = QFont("Times New Roman", 9)
                if cell.font:
                    if cell.font.name: font.setFamily(cell.font.name)
                    if cell.font.sz: font.setPointSize(int(cell.font.sz))
                    if cell.font.b: font.setBold(True)
                item.setFont(font)

                align = Qt.AlignVCenter
                if cell.alignment:
                    if cell.alignment.horizontal == 'center':
                        align |= Qt.AlignHCenter
                    elif cell.alignment.horizontal == 'right':
                        align |= Qt.AlignRight
                    else:
                        align |= Qt.AlignLeft
                    if cell.alignment.vertical == 'top': align = (align & ~Qt.AlignVCenter) | Qt.AlignTop
                item.setTextAlignment(align)

                # T≈Ço
                if cell.fill and cell.fill.patternType == 'solid':
                    fg = cell.fill.fgColor
                    if hasattr(fg, 'rgb') and fg.rgb:
                        if isinstance(fg.rgb, str):
                            hex_color = "#" + fg.rgb[2:] if len(fg.rgb) > 6 else "#" + fg.rgb
                            item.setBackground(QColor(hex_color))

                table.setItem(r - 1, c - 1, item)

                # ROTACJA
                if cell.alignment and cell.alignment.textRotation:
                    if cell.alignment.textRotation in [90, 180]:
                        table.setItemDelegateForColumn(c - 1, self.rotated_delegate)

        # Scalanie
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            table.setSpan(min_row - 1, min_col - 1,
                          max_row - min_row + 1, max_col - min_col + 1)

    # --- NOWA METODA DRUKOWANIA Z PODGLƒÑDU ---
    def print_current_file(self):
        printer_dialog = PrinterDialog(self)
        if printer_dialog.exec_() == QDialog.Accepted:
            selected_printer = printer_dialog.selected_printer
            original_printer = win32print.GetDefaultPrinter()
            try:
                win32print.SetDefaultPrinter(selected_printer)
                # Drukujemy aktualnie podglƒÖdany plik
                win32api.ShellExecute(0, "print", self.filepath, None, ".", 0)
                QMessageBox.information(self, "Sukces", f"Wys≈Çano do druku:\n{selected_printer}")
            except Exception as e:
                logging.error(f"B≈ÇƒÖd druku z podglƒÖdu: {e}")
                QMessageBox.critical(self, "B≈ÇƒÖd druku", f"WystƒÖpi≈Ç b≈ÇƒÖd: {e}")
            finally:
                win32print.SetDefaultPrinter(original_printer)


class EdytorPracownikow(QDialog):
    def __init__(self, current_text, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edytuj listƒô pracownik√≥w")
        self.resize(400, 300)
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Wpisz pracownik√≥w (Imiƒô Nazwisko, Stanowisko):"))
        self.text_edit = QTextEdit()
        self.text_edit.setPlainText(current_text)
        layout.addWidget(self.text_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText()


class KlikalnyKalendarz(QCalendarWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.custom_holidays = set()
        self.cached_holidays = holidays.PL()

    def paintCell(self, painter, rect, date):
        py_date = date.toPyDate()
        if date in self.custom_holidays:
            painter.save()
            painter.fillRect(rect, QColor("salmon"))
            painter.setPen(Qt.black)
            painter.drawText(rect, Qt.AlignCenter, str(date.day()))
            painter.restore()
        elif date.dayOfWeek() >= 6 or py_date in self.cached_holidays:
            painter.save()
            painter.fillRect(rect, QColor("#D3D3D3"))
            painter.setPen(Qt.black)
            painter.drawText(rect, Qt.AlignCenter, str(date.day()))
            painter.restore()
        else:
            super().paintCell(painter, rect, date)


# ============================================================================
# 2. LOGIKA GENEROWANIA EXCELA
# ============================================================================

class ExcelGenerator:
    def create_file(self, emp, folder, month, year, month_str, holidays_map):
        wb = Workbook()

        # --- STRONA 1 ---
        ws = wb.active
        ws.title = "Lista Obecno≈õci"
        ws.sheet_view.tabSelected = True

        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

        ws.page_margins.left = 1.78 / 2.54
        ws.page_margins.right = 1.78 / 2.54
        ws.page_margins.top = 1.91 / 2.54
        ws.page_margins.bottom = 1.91 / 2.54
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        # Style
        font_header_name = 'Cambria'
        font_body_name = 'Times New Roman'
        font_title = Font(name=font_header_name, size=14, bold=True)
        font_emp = Font(name=font_header_name, size=12, bold=True)

        font_table_header_main = Font(name=font_body_name, size=7.5, bold=True)
        font_table_header_f5 = Font(name=font_body_name, size=7, bold=False)

        font_lp_bold = Font(name=font_body_name, size=7.5, bold=True)
        font_cell = Font(name=font_body_name, size=10)
        font_small = Font(name=font_body_name, size=8)

        thin = Side(style='thin', color="000000")
        medium = Side(style='medium', color="000000")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        grey_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Kolumny
        ws.column_dimensions['A'].width = 4.3
        ws.column_dimensions['B'].width = 11.5
        ws.column_dimensions['C'].width = 11.4
        ws.column_dimensions['D'].width = 17.0
        ws.column_dimensions['E'].width = 20.5
        ws.column_dimensions['F'].width = 14.5

        # Nag≈Ç√≥wek
        ws.merge_cells('A1:F1')
        ws['A1'] = f"LISTA OBECNO≈öCI {month_str} {year} ROK"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:F4')
        ws['A2'] = f"{emp['name']}\n{emp['job']}"
        ws['A2'].alignment = align_center
        ws['A2'].font = font_emp
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15

        # Tabela Nag≈Ç√≥wki
        headers_s1 = ["Lp.", "GODZINA\nROZPOCZƒòCIA\nPRACY", "GODZINA\nZAKO≈ÉCZENIA\nPRACY", "PODPIS", "UWAGI",
                      "PODPIS OSOBY\nUPOWA≈ªNIONEJ\nDO KONTROLI"]
        ws.row_dimensions[5].height = 51.02

        for col_num, header in enumerate(headers_s1, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.border = border_thin
            if col_num == 6:
                # F5: Poziomy
                cell.font = font_table_header_f5
                cell.alignment = align_center
                cell.fill = grey_fill
            else:
                cell.font = font_table_header_main
                cell.alignment = align_center

        # Dni
        days_in_month = calendar.monthrange(year, month)[1]
        row_height_points_s1 = 18.2

        for day in range(1, 32):
            row = 5 + day
            ws.row_dimensions[row].height = row_height_points_s1
            cell_lp = ws.cell(row=row, column=1)
            cell_lp.alignment = align_center
            cell_lp.border = border_thin
            cell_lp.font = font_lp_bold
            for col in range(2, 7):
                c = ws.cell(row=row, column=col)
                c.border = border_thin
                c.font = font_cell
            if day <= days_in_month:
                cell_lp.value = day
                if QDate(year, month, day) in holidays_map:
                    for col in range(1, 7): ws.cell(row=row, column=col).fill = grey_fill
            else:
                cell_lp.value = "X"
                for col in range(2, 7):
                    ws.cell(row=row, column=col).value = "X"
                    ws.cell(row=row, column=col).alignment = align_center

        # Stopka
        last_row_s1 = 37
        ws.row_dimensions[last_row_s1].height = 20
        ws.merge_cells(f'A{last_row_s1}:E{last_row_s1}')
        ws[f'A{last_row_s1}'] = "Oznaczenia: N - nieobecno≈õƒá;"
        ws[f'A{last_row_s1}'].font = font_small
        ws[f'A{last_row_s1}'].alignment = Alignment(horizontal='left', vertical='center')

        ws.cell(row=last_row_s1, column=6).border = Border(top=medium)

        # Ramki Grube
        for col in range(1, 7):
            ws.cell(row=1, column=col).border = Border(top=medium, bottom=ws.cell(row=1, column=col).border.bottom,
                                                       left=ws.cell(row=1, column=col).border.left,
                                                       right=ws.cell(row=1, column=col).border.right)
        for col in range(1, 7):
            ws.cell(row=36, column=col).border = Border(top=ws.cell(row=36, column=col).border.top, bottom=medium,
                                                        left=ws.cell(row=36, column=col).border.left,
                                                        right=ws.cell(row=36, column=col).border.right)
        for row in range(1, 37):
            ws.cell(row=row, column=1).border = Border(top=ws.cell(row=row, column=1).border.top,
                                                       bottom=ws.cell(row=row, column=1).border.bottom, left=medium,
                                                       right=ws.cell(row=row, column=1).border.right)
        for row in range(1, 37):
            ws.cell(row=row, column=6).border = Border(top=ws.cell(row=row, column=6).border.top,
                                                       bottom=ws.cell(row=row, column=6).border.bottom,
                                                       left=ws.cell(row=row, column=6).border.left, right=medium)

        # --- STRONA 2 ---
        ws2 = wb.create_sheet("Ewidencja")
        ws2.sheet_view.tabSelected = True

        ws2.page_setup.paperSize = 9
        ws2.page_setup.orientation = 'portrait'
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 1
        ws2.print_options.horizontalCentered = True

        ws2.page_margins.left = 1.78 / 2.54
        ws2.page_margins.right = 1.78 / 2.54
        ws2.page_margins.top = 1.91 / 2.54
        ws2.page_margins.bottom = 1.50 / 2.54

        font_s2_group_header = Font(name=font_body_name, size=8, bold=False)
        font_s2_vert_8 = Font(name=font_body_name, size=8, bold=False)
        font_s2_vert_7 = Font(name=font_body_name, size=7, bold=False)
        font_s2_day_lp = Font(name=font_body_name, size=10, bold=False)
        font_s2_cell_data = Font(name=font_body_name, size=10, bold=False)
        font_s2_razem = Font(name=font_body_name, size=7, bold=False)
        font_s2_legend = Font(name=font_body_name, size=8, bold=False)
        font_s2_sign = Font(name=font_body_name, size=10, bold=False)
        align_vertical_90 = Alignment(textRotation=90, horizontal='center', vertical='center', wrap_text=True)

        ws2.column_dimensions['A'].width = 3.6
        ws2.column_dimensions['B'].width = 6.0
        for col_letter in ['C', 'D', 'E']: ws2.column_dimensions[col_letter].width = 4.2
        for col_letter in ['F', 'G', 'H', 'I']: ws2.column_dimensions[col_letter].width = 4.2
        for col_letter in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']: ws2.column_dimensions[col_letter].width = 3.3
        ws2.column_dimensions['R'].width = 9.5

        ws2.row_dimensions[1].height = 31
        ws2.row_dimensions[2].height = 76

        ws2.merge_cells('A1:A2')
        ws2['A1'] = "Dzie≈Ñ\nmiesiƒÖca"
        ws2['A1'].font = font_s2_vert_8
        ws2['A1'].alignment = align_vertical_90
        ws2['A1'].border = border_thin

        ws2['B1'] = "Czas\npracy"
        ws2['B1'].font = font_s2_group_header
        ws2['B1'].alignment = align_center
        ws2['B1'].border = border_thin

        ws2.merge_cells('C1:E1')
        ws2['C1'] = "Czas przepracowany w\ngodzinach"
        ws2['C1'].font = font_s2_group_header
        ws2['C1'].alignment = align_center
        ws2['C1'].border = border_thin

        ws2.merge_cells('F1:Q1')
        ws2['F1'] = "Czas nieobecno≈õci w pracy w godzinach"
        ws2['F1'].font = font_s2_group_header
        ws2['F1'].alignment = align_center
        ws2['F1'].border = border_thin

        ws2['R1'] = "Normatywny\nczas pracy z\nKP"
        ws2['R1'].font = font_s2_group_header
        ws2['R1'].alignment = align_center
        ws2['R1'].border = border_thin
        ws2['R2'].border = border_thin

        # Obliczanie normy (R2)
        working_days_kp = 0
        for day in range(1, days_in_month + 1):
            q_date = QDate(year, month, day)
            if q_date in holidays_map: continue
            working_days_kp += 1
        working_hours_kp = working_days_kp * 8
        ws2['R2'].value = f"w dniach: {working_days_kp}\nw godzinach: {working_hours_kp}"
        ws2['R2'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws2['R2'].font = font_s2_group_header

        headers_row2 = {
            'B': "Liczba godzin\nprzepracowanych",
            'C': "Niedziele i ≈õwiƒôta",
            'D': "W porze nocnej",
            'E': "W godzinach\nnadliczbowych",
            'F': "Urlop wypoczynkowy",
            'G': "Opieka KP 188 ¬ß1",
            'H': "Opieka zasi≈Çek",
            'I': "L4 zwolnienie lekarskie",
            'J': "Urlop okoliczno≈õciowy",
            'K': "Urlop wypoczynkowy\n\"na ≈ºƒÖdanie\"",
            'L': "Urlop bezp≈Çatny",
            'M': "Dni wolne za ≈õwiƒôto\nprzypadajƒÖce w sobotƒô",
            'N': "Urlop macierzy≈Ñski",
            'O': "Urlop rodzicielski",
            'P': "Urlop dodatkowy",
            'Q': ""
        }

        for col_idx, col_char in enumerate("ABCDEFGHIJKLMNOPQR", 1):
            if col_idx == 1 or col_idx == 18: continue
            cell = ws2.cell(row=2, column=col_idx)
            cell.border = border_thin
            if col_char in headers_row2:
                cell.value = headers_row2[col_char]
                cell.alignment = align_vertical_90
                if col_idx <= 5:
                    cell.font = font_s2_vert_8
                else:
                    cell.font = font_s2_vert_7
            if col_idx >= 3 and col_idx <= 17:
                ws2.cell(row=1, column=col_idx).border = border_thin

        # Dane
        row_height_s2 = 17
        for day in range(1, 32):
            row = 2 + day
            ws2.row_dimensions[row].height = row_height_s2
            cell_day = ws2.cell(row=row, column=1)
            cell_day.alignment = align_center
            cell_day.border = border_thin
            cell_day.font = font_s2_day_lp
            for col in range(2, 19):
                c = ws2.cell(row=row, column=col)
                c.border = border_thin
                c.font = font_s2_cell_data
                c.alignment = align_center
            if day <= days_in_month:
                cell_day.value = day
                if QDate(year, month, day) in holidays_map:
                    for col in range(1, 19): ws2.cell(row=row, column=col).fill = grey_fill
            else:
                cell_day.value = "X"
                for col in range(2, 19): ws2.cell(row=row, column=col).value = "X"

        row_razem = 34
        ws2.row_dimensions[row_razem].height = row_height_s2
        cell_razem = ws2.cell(row=row_razem, column=1, value="razem")
        cell_razem.font = font_s2_razem
        cell_razem.alignment = align_center
        cell_razem.border = border_thin
        for col in range(2, 19):
            c = ws2.cell(row=row_razem, column=col)
            c.border = border_thin
            c.font = font_s2_razem

        footer_row = 35
        ws2.row_dimensions[footer_row].height = 77.4
        ws2.merge_cells(f'A{footer_row}:R{footer_row}')

        # --- PRZYWR√ìCONA LEGENDA ---
        legend_text = (
            "Urlop wypoczynkowy, Opieka Art 188¬ß1, Opieka zasi≈Çek, L4, Urlop okoliczno≈õciowy, Urlop wypoczynkowy \"na ≈ºƒÖdanie\",  Urlop bezp≈Çatny,  "
            "Dn - Wyp≈Çata nadgodzin, W≈õ - Wolne za pracƒô w ≈õwiƒôto, Wn - Wolne za nadgodziny, W5 - Wolne z tytu≈Çu 5-dniowego tygodnia pracy, "
            "W - Urlop wychowawczy, Um - Urlop macierzy≈Ñski, ≈öR - ≈öwiadczenia reabilitacyjne, ≈ös - ≈öwiadek w sƒÖdzie, Kr - Oddanie krwi., "
            "S-szkolenie z udzielania dziecku pierwszej pomocy, D-delegacje/podr√≥≈ºe s≈Çu≈ºbowe, P-dni wolne na poszukiwanie pracy, "
            "Urek-nieobecno≈õci za kt√≥re zgodnie z par. 16 pkt 2 rozporzƒÖdzenia w sprawie sposobu usprawiedliwienia nieobecno≈õci w pracy "
            "oraz udzielania pracownikom zwolnie≈Ñ od pracy, pracownicy majƒÖ prawo do uzyskania rekompensaty pieniƒô≈ºnej od w≈Ça≈õciwego organu,"
            "Sn - sp√≥≈∫nienie, N-nieusprawiedliwione nieobecno≈õci w pracy, Nun-nieobecno≈õƒá usprawiedliwiona niep≈Çatna"
        )
        leg_cell = ws2[f'A{footer_row}']
        leg_cell.value = legend_text
        leg_cell.font = font_s2_legend
        leg_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        leg_cell.border = border_thin

        sign_row = footer_row + 1
        ws2.row_dimensions[sign_row].height = 13
        ws2.merge_cells(f'A{sign_row}:E{sign_row}')
        ws2[f'A{sign_row}'] = "PODPIS KADR"
        ws2[f'A{sign_row}'].alignment = align_center
        ws2[f'A{sign_row}'].font = font_s2_sign
        ws2[f'A{sign_row}'].border = None

        ws2.merge_cells(f'F{sign_row}:R{sign_row}')
        ws2[f'F{sign_row}'] = "PODPIS DYREKTORA ≈ª≈ÅOBKA"
        ws2[f'F{sign_row}'].alignment = align_center
        ws2[f'F{sign_row}'].font = font_s2_sign
        ws2[f'F{sign_row}'].border = None

        frame_last_row = 35
        for col in range(1, 19):
            ws2.cell(row=1, column=col).border = Border(top=medium, bottom=thin, left=thin, right=thin)
        ws2['A35'].border = Border(top=thin, bottom=medium, left=medium, right=medium)
        for row in range(1, frame_last_row + 1):
            c = ws2.cell(row=row, column=1)
            curr = c.border
            c.border = Border(top=curr.top, bottom=curr.bottom, left=medium, right=curr.right)
        for row in range(1, frame_last_row + 1):
            c = ws2.cell(row=row, column=18)
            curr = c.border
            c.border = Border(top=curr.top, bottom=curr.bottom, left=curr.left, right=medium)

        filename = f"{emp['name'].replace(' ', '_')}_{month_str}_{year}.xlsx"
        path = os.path.join(folder, filename)
        wb.save(path)
        return path


class WorkerThread(QThread):
    progress_updated = pyqtSignal(int)
    finished = pyqtSignal(dict, str)
    error_occurred = pyqtSignal(str)

    def __init__(self, employees, folder, month, year, month_str, holidays_map):
        super().__init__()
        self.employees = employees
        self.folder = folder
        self.month = month
        self.year = year
        self.month_str = month_str
        self.holidays_map = holidays_map
        self.generator = ExcelGenerator()

    def run(self):
        generated_map = {}
        total = len(self.employees)
        try:
            for i, emp in enumerate(self.employees):
                # Je≈õli folder nie jest zdefiniowany (np. u≈ºytkownik anulowa≈Ç), wƒÖtek siƒô nie uda
                if not self.folder:
                    raise Exception("Nie wybrano folderu zapisu.")

                filepath = self.generator.create_file(emp, self.folder, self.month, self.year, self.month_str,
                                                      self.holidays_map)
                generated_map[emp['key']] = filepath
                progress_percent = int(((i + 1) / total) * 100)
                self.progress_updated.emit(progress_percent)

            self.finished.emit(generated_map, self.folder)

        except PermissionError as e:
            clean_msg = (f"Nie mo≈ºna zapisaƒá pliku dla pracownika: {emp['name']}.\n\n"
                         f"PRAWDOPODOBNA PRZYCZYNA:\n"
                         f"Plik jest otwarty w Excelu/OpenOffice.\n\n"
                         f"Zamknij plik i spr√≥buj ponownie.")
            logging.error(f"PermissionError: {e}")
            self.error_occurred.emit(clean_msg)

        except Exception as e:
            logging.error(f"Critical Error: {e}")
            self.error_occurred.emit(f"WystƒÖpi≈Ç nieoczekiwany b≈ÇƒÖd:\n{str(e)}")


# ==========================================
# 4. G≈Å√ìWNA APLIKACJA
# ==========================================

class GeneratorListObecnosci(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generator List Obecno≈õci v20.0 (Final Complete)")
        self.setGeometry(100, 100, 1100, 800)
        self.generated_files_map = {}

        self.config = configparser.ConfigParser()
        self.last_save_folder = ""
        self.load_config()

        self.initUI()
        self.load_employees_on_startup()

    def load_config(self):
        if os.path.exists('config.ini'):
            try:
                self.config.read('config.ini')
                self.last_save_folder = self.config.get('SETTINGS', 'LastFolder', fallback="")
            except Exception:
                pass

    def save_config(self):
        if 'SETTINGS' not in self.config:
            self.config['SETTINGS'] = {}
        self.config['SETTINGS']['LastFolder'] = self.last_save_folder
        with open('config.ini', 'w') as configfile:
            self.config.write(configfile)

    def initUI(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout()

        # LEWY
        left_panel = QVBoxLayout()

        date_group = QGroupBox("1. Wybierz okres")
        date_layout = QVBoxLayout()
        self.month_cb = QComboBox()
        self.month_names = ["Stycze≈Ñ", "Luty", "Marzec", "Kwiecie≈Ñ", "Maj", "Czerwiec",
                            "Lipiec", "Sierpie≈Ñ", "Wrzesie≈Ñ", "Pa≈∫dziernik", "Listopad", "Grudzie≈Ñ"]
        self.month_cb.addItems(self.month_names)
        self.year_cb = QComboBox()
        current_year = QDate.currentDate().year()
        for y in range(current_year - 1, current_year + 5):
            self.year_cb.addItem(str(y))
        self.year_cb.setCurrentText(str(current_year))
        self.month_cb.currentIndexChanged.connect(self.sync_calendar_from_combo)
        self.year_cb.currentIndexChanged.connect(self.sync_calendar_from_combo)
        date_layout.addWidget(QLabel("MiesiƒÖc:"))
        date_layout.addWidget(self.month_cb)
        date_layout.addWidget(QLabel("Rok:"))
        date_layout.addWidget(self.year_cb)
        date_group.setLayout(date_layout)
        left_panel.addWidget(date_group)

        emp_group = QGroupBox("2. Pracownicy")
        emp_layout = QVBoxLayout()
        self.emp_list_widget = QListWidget()
        self.emp_list_widget.setSelectionMode(QAbstractItemView.NoSelection)
        self.emp_list_widget.itemChanged.connect(self.update_print_button_state)
        # PODW√ìJNE KLIKNIƒòCIE OTWIERA PODGLƒÑD WBUDOWANY
        self.emp_list_widget.itemDoubleClicked.connect(self.open_preview_window)
        emp_layout.addWidget(self.emp_list_widget)
        edit_btns_layout = QHBoxLayout()
        self.btn_edit_text = QPushButton("Edytuj listƒô")
        self.btn_edit_text.clicked.connect(self.open_text_editor)
        self.btn_toggle_select = QPushButton("Odznacz wszystkich")
        self.btn_toggle_select.clicked.connect(self.toggle_selection)
        edit_btns_layout.addWidget(self.btn_edit_text)
        edit_btns_layout.addWidget(self.btn_toggle_select)
        emp_layout.addLayout(edit_btns_layout)
        emp_group.setLayout(emp_layout)
        left_panel.addWidget(emp_group)

        action_group = QGroupBox("4. Akcje")
        action_layout = QVBoxLayout()
        self.generate_btn = QPushButton("Generuj wybrane")
        self.generate_btn.setMinimumHeight(40)
        self.generate_btn.setStyleSheet("font-weight: bold; font-size: 14px; background-color: #4CAF50; color: white;")
        self.generate_btn.clicked.connect(self.start_generation_thread)
        action_layout.addWidget(self.generate_btn)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        action_layout.addWidget(self.progress_bar)

        self.print_btn = QPushButton("Drukuj karty...")
        self.print_btn.setMinimumHeight(40)
        self.print_btn.setStyleSheet("font-weight: bold; font-size: 12px; background-color: #2196F3; color: white;")
        self.print_btn.clicked.connect(self.print_generated_files)
        self.print_btn.setEnabled(False)
        action_layout.addWidget(self.print_btn)

        # PRZYCISKI FOLDERU
        folder_btns_layout = QHBoxLayout()

        self.open_folder_btn = QPushButton("Otw√≥rz folder")
        self.open_folder_btn.clicked.connect(self.open_last_folder)
        self.open_folder_btn.setEnabled(False)

        self.change_folder_btn = QPushButton("Zmie≈Ñ folder zapisu")
        self.change_folder_btn.clicked.connect(self.change_save_folder)

        folder_btns_layout.addWidget(self.open_folder_btn)
        folder_btns_layout.addWidget(self.change_folder_btn)
        action_layout.addLayout(folder_btns_layout)

        # Je≈õli folder jest w configu, aktywujemy przycisk
        if self.last_save_folder and os.path.exists(self.last_save_folder):
            self.open_folder_btn.setEnabled(True)

        action_group.setLayout(action_layout)
        left_panel.addWidget(action_group)
        main_layout.addLayout(left_panel, stretch=1)

        # PRAWY
        right_panel = QVBoxLayout()
        cal_group = QGroupBox("3. Kalendarz (≈öwiƒôta i dni wolne)")
        cal_layout = QVBoxLayout()
        cal_layout.addWidget(QLabel(
            "SZARY = Wolne ustawowo | CZERWONY = Wolne dodatkowe\nU≈ºyj rolki myszy na kalendarzu, aby zmieniƒá miesiƒÖc."))
        self.calendar = KlikalnyKalendarz()
        self.calendar.setGridVisible(True)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        self.calendar.setNavigationBarVisible(False)
        self.calendar.currentPageChanged.connect(self.sync_combo_from_calendar)
        self.calendar.clicked.connect(self.toggle_holiday)
        cal_layout.addWidget(self.calendar)
        cal_group.setLayout(cal_layout)
        right_panel.addWidget(cal_group, stretch=2)
        main_layout.addLayout(right_panel)
        main_widget.setLayout(main_layout)

        current_month_idx = QDate.currentDate().month() - 1
        self.month_cb.setCurrentIndex(current_month_idx)
        self.sync_calendar_from_combo()

    def sync_calendar_from_combo(self):
        self.calendar.blockSignals(True)
        month = self.month_cb.currentIndex() + 1
        year = int(self.year_cb.currentText())
        self.calendar.setCurrentPage(year, month)
        self.calendar.cached_holidays = holidays.PL(years=year)
        self.calendar.blockSignals(False)

    def sync_combo_from_calendar(self, year, month):
        self.month_cb.blockSignals(True)
        self.year_cb.blockSignals(True)
        self.year_cb.setCurrentText(str(year))
        self.month_cb.setCurrentIndex(month - 1)
        self.calendar.cached_holidays = holidays.PL(years=year)
        self.month_cb.blockSignals(False)
        self.year_cb.blockSignals(False)

    def load_employees_on_startup(self):
        filename = "pracownicy.txt"
        if not os.path.exists(filename):
            try:
                with open(filename, "w", encoding="utf-8") as f:
                    f.write("Jan Kowalski, Opiekun\nAnna Nowak, Kucharka")
            except Exception:
                pass
        try:
            with open(filename, "r", encoding="utf-8") as f:
                content = f.read()
                self.update_list_widget(content)
        except Exception as e:
            QMessageBox.warning(self, "B≈ÇƒÖd", f"Problem z plikiem pracownicy.txt: {e}")

    def update_list_widget(self, text):
        self.emp_list_widget.blockSignals(True)
        self.emp_list_widget.clear()
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if line:
                item = QListWidgetItem(line)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Checked)
                self.emp_list_widget.addItem(item)
        self.emp_list_widget.blockSignals(False)
        self.btn_toggle_select.setText("Odznacz wszystkich")

    def open_text_editor(self):
        current_text = ""
        for i in range(self.emp_list_widget.count()):
            current_text += self.emp_list_widget.item(i).text() + "\n"
        dialog = EdytorPracownikow(current_text, self)
        if dialog.exec_():
            new_text = dialog.get_text()
            self.update_list_widget(new_text)
            try:
                with open("pracownicy.txt", "w", encoding="utf-8") as f:
                    f.write(new_text)
            except Exception as e:
                QMessageBox.warning(self, "B≈ÇƒÖd zapisu", str(e))

    def toggle_selection(self):
        self.emp_list_widget.blockSignals(True)
        current_text = self.btn_toggle_select.text()
        if current_text == "Odznacz wszystkich":
            for i in range(self.emp_list_widget.count()):
                self.emp_list_widget.item(i).setCheckState(Qt.Unchecked)
            self.btn_toggle_select.setText("Zaznacz wszystkich")
        else:
            for i in range(self.emp_list_widget.count()):
                self.emp_list_widget.item(i).setCheckState(Qt.Checked)
            self.btn_toggle_select.setText("Odznacz wszystkich")
        self.emp_list_widget.blockSignals(False)
        self.update_print_button_state()

    def toggle_holiday(self, date):
        holidays_set = self.calendar.custom_holidays
        if date in holidays_set:
            holidays_set.remove(date)
        else:
            holidays_set.add(date)
        self.calendar.updateCell(date)

    def calculate_holidays(self, year, month):
        final_holidays = set()
        pl_holidays = holidays.PL(years=year)
        days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, days_in_month + 1):
            current_date = QDate(year, month, day)
            py_date = current_date.toPyDate()
            if current_date.dayOfWeek() >= 6:
                final_holidays.add(current_date)
            elif py_date in pl_holidays:
                final_holidays.add(current_date)
            elif current_date in self.calendar.custom_holidays:
                final_holidays.add(current_date)
        return final_holidays

    def update_print_button_state(self):
        count = 0
        for i in range(self.emp_list_widget.count()):
            item = self.emp_list_widget.item(i)
            if item.checkState() == Qt.Checked and item.text() in self.generated_files_map:
                count += 1

        if count > 0:
            self.print_btn.setEnabled(True)
            self.print_btn.setText(f"Drukuj {count} kart...")
        else:
            if not self.generated_files_map:
                self.print_btn.setEnabled(False)
                self.print_btn.setText("Drukuj karty...")
            else:
                self.print_btn.setEnabled(False)
                self.print_btn.setText("Zaznacz pracownik√≥w do druku")

    # --- ZMIANA: ZMIANA FOLDERU ZAPISU ---
    def change_save_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Wybierz nowy folder domy≈õlny")
        if folder:
            self.last_save_folder = folder
            self.save_config()
            self.open_folder_btn.setEnabled(True)
            QMessageBox.information(self, "Sukces", f"Domy≈õlny folder zmieniony na:\n{folder}")

    def start_generation_thread(self):
        employees_to_generate = []
        for i in range(self.emp_list_widget.count()):
            item = self.emp_list_widget.item(i)
            if item.checkState() == Qt.Checked:
                line = item.text()
                parts = line.split(',')
                name = parts[0].strip()
                job = parts[1].strip() if len(parts) > 1 else "Pracownik"
                employees_to_generate.append({'name': name, 'job': job, 'key': line})

        if not employees_to_generate:
            QMessageBox.warning(self, "B≈ÇƒÖd", "Nie zaznaczono ≈ºadnego pracownika!")
            return

        # Ustalanie folderu docelowego
        target_folder = ""

        # 1. Sprawdzamy, czy jest zapisany w configu
        if self.last_save_folder and os.path.exists(self.last_save_folder):
            target_folder = self.last_save_folder
        else:
            # 2. Je≈õli nie ma, tworzymy folder domy≈õlny obok programu
            default_path = os.path.join(os.getcwd(), "Wygenerowane Karty")
            if not os.path.exists(default_path):
                try:
                    os.makedirs(default_path)
                except Exception:
                    pass
            target_folder = default_path

            # Zapisujemy ten domy≈õlny folder jako ostatni u≈ºywany
            self.last_save_folder = target_folder
            self.save_config()
            self.open_folder_btn.setEnabled(True)

        month_idx = self.month_cb.currentIndex() + 1
        year = int(self.year_cb.currentText())
        month_name = self.month_names[month_idx - 1].upper()
        holidays_map = self.calculate_holidays(year, month_idx)

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.generate_btn.setEnabled(False)
        self.print_btn.setEnabled(False)

        self.thread = WorkerThread(employees_to_generate, target_folder, month_idx, year, month_name, holidays_map)
        self.thread.progress_updated.connect(self.progress_bar.setValue)
        self.thread.finished.connect(self.on_generation_finished)
        self.thread.error_occurred.connect(self.on_generation_error)
        self.thread.start()

    def on_generation_finished(self, generated_map, folder):
        self.generated_files_map = generated_map
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        self.update_print_button_state()
        QMessageBox.information(self, "Sukces", f"Wygenerowano plik√≥w: {len(generated_map)}\nLokalizacja: {folder}")

    def on_generation_error(self, error_msg):
        self.progress_bar.setVisible(False)
        self.generate_btn.setEnabled(True)
        QMessageBox.critical(self, "B≈ÇƒÖd generowania", error_msg)

    def open_last_folder(self):
        if self.last_save_folder and os.path.exists(self.last_save_folder):
            os.startfile(self.last_save_folder)
        else:
            QMessageBox.warning(self, "Info", "Folder nie zosta≈Ç jeszcze wybrany lub nie istnieje.")

    def open_preview_window(self, item):
        employee_key = item.text()
        if employee_key in self.generated_files_map:
            filepath = self.generated_files_map[employee_key]
            if os.path.exists(filepath):
                # TO OTWIERA WBUDOWANY PODGLƒÑD (NOWY DLA WERSJI 20.0)
                preview = ExcelPreviewDialog(filepath, self)
                preview.exec_()
            else:
                QMessageBox.warning(self, "B≈ÇƒÖd", "Plik nie istnieje.")
        else:
            QMessageBox.information(self, "Info", "Najpierw wygeneruj karty.")

    def print_generated_files(self):
        files_to_print = []
        for i in range(self.emp_list_widget.count()):
            item = self.emp_list_widget.item(i)
            if item.checkState() == Qt.Checked and item.text() in self.generated_files_map:
                files_to_print.append(self.generated_files_map[item.text()])

        if not files_to_print:
            QMessageBox.warning(self, "Info", "Brak zaznaczonych kart do wydruku.")
            return

        printer_dialog = PrinterDialog(self)
        if printer_dialog.exec_() == QDialog.Accepted:
            selected_printer = printer_dialog.selected_printer
            original_printer = win32print.GetDefaultPrinter()
            try:
                win32print.SetDefaultPrinter(selected_printer)
                for filepath in files_to_print:
                    win32api.ShellExecute(0, "print", filepath, None, ".", 0)
                    time.sleep(2)
                QMessageBox.information(self, "Sukces", f"Wys≈Çano {len(files_to_print)} plik√≥w do: {selected_printer}")
            except Exception as e:
                logging.error(f"B≈ÇƒÖd druku: {e}")
                QMessageBox.critical(self, "B≈ÇƒÖd druku", f"WystƒÖpi≈Ç b≈ÇƒÖd: {e}")
            finally:
                win32print.SetDefaultPrinter(original_printer)


if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        ex = GeneratorListObecnosci()
        ex.show()
        sys.exit(app.exec_())
    except Exception as e:
        logging.critical(f"Krytyczny b≈ÇƒÖd aplikacji: {e}")