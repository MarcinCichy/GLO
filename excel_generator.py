import os
import calendar

from PyQt5.QtCore import QDate

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelGenerator:
    def create_file(self, emp, folder, month, year, month_str, holidays_map):
        wb = Workbook()

        # --- STRONA 1 ---
        ws = wb.active
        ws.title = "Lista Obecności"
        ws.sheet_view.tabSelected = True

        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

        ws.page_margins.left = 1.78 / 2.54
        ws.page_margins.right = 1.78 / 2.54
        ws.page_margins.top = 1.91 / 2.54
        ws.page_margins.bottom = 1.91 / 2.54
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        # Style
        font_header_name = 'Cambria'
        font_body_name = 'Times New Roman'
        font_title = Font(name=font_header_name, size=14, bold=True)
        font_emp = Font(name=font_header_name, size=12, bold=True)

        font_table_header_main = Font(name=font_body_name, size=7.5, bold=True)
        font_table_header_f5 = Font(name=font_body_name, size=7, bold=False)

        font_lp_bold = Font(name=font_body_name, size=7.5, bold=True)
        font_cell = Font(name=font_body_name, size=10)
        font_small = Font(name=font_body_name, size=8)

        thin = Side(style='thin', color="000000")
        medium = Side(style='medium', color="000000")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        grey_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Kolumny
        ws.column_dimensions['A'].width = 4.3
        ws.column_dimensions['B'].width = 11.5
        ws.column_dimensions['C'].width = 11.4
        ws.column_dimensions['D'].width = 17.0
        ws.column_dimensions['E'].width = 20.5
        ws.column_dimensions['F'].width = 14.5

        # Nagłówek
        ws.merge_cells('A1:F1')
        ws['A1'] = f"LISTA OBECNOŚCI {month_str} {year} ROK"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 25

        ws.merge_cells('A2:F4')
        ws['A2'] = f"{emp['name']}\n{emp['job']}"
        ws['A2'].alignment = align_center
        ws['A2'].font = font_emp
        ws.row_dimensions[2].height = 15
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15

        # Tabela Nagłówki
        headers_s1 = ["Lp.", "GODZINA\nROZPOCZĘCIA\nPRACY", "GODZINA\nZAKOŃCZENIA\nPRACY", "PODPIS", "UWAGI",
                      "PODPIS OSOBY\nUPOWAŻNIONEJ\nDO KONTROLI"]
        ws.row_dimensions[5].height = 51.02

        for col_num, header in enumerate(headers_s1, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.border = border_thin
            if col_num == 6:
                # F5: Poziomy
                cell.font = font_table_header_f5
                cell.alignment = align_center
                cell.fill = grey_fill
            else:
                cell.font = font_table_header_main
                cell.alignment = align_center

        # Dni
        days_in_month = calendar.monthrange(year, month)[1]
        row_height_points_s1 = 18.2

        for day in range(1, 32):
            row = 5 + day
            ws.row_dimensions[row].height = row_height_points_s1
            cell_lp = ws.cell(row=row, column=1)
            cell_lp.alignment = align_center
            cell_lp.border = border_thin
            cell_lp.font = font_lp_bold
            for col in range(2, 7):
                c = ws.cell(row=row, column=col)
                c.border = border_thin
                c.font = font_cell
            if day <= days_in_month:
                cell_lp.value = day
                if QDate(year, month, day) in holidays_map:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = grey_fill
            else:
                cell_lp.value = "X"
                for col in range(2, 7):
                    ws.cell(row=row, column=col).value = "X"
                    ws.cell(row=row, column=col).alignment = align_center

        # Stopka
        last_row_s1 = 37
        ws.row_dimensions[last_row_s1].height = 20
        ws.merge_cells(f'A{last_row_s1}:E{last_row_s1}')
        ws[f'A{last_row_s1}'] = "Oznaczenia: N - nieobecność;"
        ws[f'A{last_row_s1}'].font = font_small
        ws[f'A{last_row_s1}'].alignment = Alignment(horizontal='left', vertical='center')

        ws.cell(row=last_row_s1, column=6).border = Border(top=medium)

        # Ramki Grube
        for col in range(1, 7):
            ws.cell(row=1, column=col).border = Border(top=medium, bottom=ws.cell(row=1, column=col).border.bottom,
                                                       left=ws.cell(row=1, column=col).border.left,
                                                       right=ws.cell(row=1, column=col).border.right)
        for col in range(1, 7):
            ws.cell(row=36, column=col).border = Border(top=ws.cell(row=36, column=col).border.top, bottom=medium,
                                                        left=ws.cell(row=36, column=col).border.left,
                                                        right=ws.cell(row=36, column=col).border.right)
        for row in range(1, 37):
            ws.cell(row=row, column=1).border = Border(top=ws.cell(row=row, column=1).border.top,
                                                       bottom=ws.cell(row=row, column=1).border.bottom, left=medium,
                                                       right=ws.cell(row=row, column=1).border.right)
        for row in range(1, 37):
            ws.cell(row=row, column=6).border = Border(top=ws.cell(row=row, column=6).border.top,
                                                       bottom=ws.cell(row=row, column=6).border.bottom,
                                                       left=ws.cell(row=row, column=6).border.left, right=medium)

        # --- STRONA 2 ---
        ws2 = wb.create_sheet("Ewidencja")
        ws2.sheet_view.tabSelected = True

        ws2.page_setup.paperSize = 9
        ws2.page_setup.orientation = 'portrait'
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 1
        ws2.print_options.horizontalCentered = True

        ws2.page_margins.left = 1.78 / 2.54
        ws2.page_margins.right = 1.78 / 2.54
        ws2.page_margins.top = 1.91 / 2.54
        ws2.page_margins.bottom = 1.50 / 2.54

        font_s2_group_header = Font(name=font_body_name, size=8, bold=False)
        font_s2_vert_8 = Font(name=font_body_name, size=8, bold=False)
        font_s2_vert_7 = Font(name=font_body_name, size=7, bold=False)
        font_s2_day_lp = Font(name=font_body_name, size=10, bold=False)
        font_s2_cell_data = Font(name=font_body_name, size=10, bold=False)
        font_s2_razem = Font(name=font_body_name, size=7, bold=False)
        font_s2_legend = Font(name=font_body_name, size=8, bold=False)
        font_s2_sign = Font(name=font_body_name, size=10, bold=False)
        align_vertical_90 = Alignment(textRotation=90, horizontal='center', vertical='center', wrap_text=True)

        ws2.column_dimensions['A'].width = 3.6
        ws2.column_dimensions['B'].width = 6.0
        for col_letter in ['C', 'D', 'E']:
            ws2.column_dimensions[col_letter].width = 4.2
        for col_letter in ['F', 'G', 'H', 'I']:
            ws2.column_dimensions[col_letter].width = 4.2
        for col_letter in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
            ws2.column_dimensions[col_letter].width = 3.3
        ws2.column_dimensions['R'].width = 9.5

        ws2.row_dimensions[1].height = 31
        ws2.row_dimensions[2].height = 76

        ws2.merge_cells('A1:A2')
        ws2['A1'] = "Dzień\nmiesiąca"
        ws2['A1'].font = font_s2_vert_8
        ws2['A1'].alignment = align_vertical_90
        ws2['A1'].border = border_thin

        ws2['B1'] = "Czas\npracy"
        ws2['B1'].font = font_s2_group_header
        ws2['B1'].alignment = align_center
        ws2['B1'].border = border_thin

        ws2.merge_cells('C1:E1')
        ws2['C1'] = "Czas przepracowany w\ngodzinach"
        ws2['C1'].font = font_s2_group_header
        ws2['C1'].alignment = align_center
        ws2['C1'].border = border_thin

        ws2.merge_cells('F1:Q1')
        ws2['F1'] = "Czas nieobecności w pracy w godzinach"
        ws2['F1'].font = font_s2_group_header
        ws2['F1'].alignment = align_center
        ws2['F1'].border = border_thin

        ws2['R1'] = "Normatywny\nczas pracy z\nKP"
        ws2['R1'].font = font_s2_group_header
        ws2['R1'].alignment = align_center
        ws2['R1'].border = border_thin
        ws2['R2'].border = border_thin

        # Obliczanie normy (R2)
        working_days_kp = 0
        for day in range(1, days_in_month + 1):
            q_date = QDate(year, month, day)
            if q_date in holidays_map:
                continue
            working_days_kp += 1
        working_hours_kp = working_days_kp * 8
        ws2['R2'].value = f"w dniach: {working_days_kp}\nw godzinach: {working_hours_kp}"
        ws2['R2'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws2['R2'].font = font_s2_group_header

        headers_row2 = {
            'B': "Liczba godzin\nprzepracowanych",
            'C': "Niedziele i święta",
            'D': "W porze nocnej",
            'E': "W godzinach\nnadliczbowych",
            'F': "Urlop wypoczynkowy",
            'G': "Opieka KP 188 §1",
            'H': "Opieka zasiłek",
            'I': "L4 zwolnienie lekarskie",
            'J': "Urlop okolicznościowy",
            'K': "Urlop wypoczynkowy\n\"na żądanie\"",
            'L': "Urlop bezpłatny",
            'M': "Dni wolne za święto\nprzypadające w sobotę",
            'N': "Urlop macierzyński",
            'O': "Urlop rodzicielski",
            'P': "Urlop dodatkowy",
            'Q': ""
        }

        for col_idx, col_char in enumerate("ABCDEFGHIJKLMNOPQR", 1):
            if col_idx == 1 or col_idx == 18:
                continue
            cell = ws2.cell(row=2, column=col_idx)
            cell.border = border_thin
            if col_char in headers_row2:
                cell.value = headers_row2[col_char]
                cell.alignment = align_vertical_90
                if col_idx <= 5:
                    cell.font = font_s2_vert_8
                else:
                    cell.font = font_s2_vert_7
            if col_idx >= 3 and col_idx <= 17:
                ws2.cell(row=1, column=col_idx).border = border_thin

        # Dane
        row_height_s2 = 17
        for day in range(1, 32):
            row = 2 + day
            ws2.row_dimensions[row].height = row_height_s2
            cell_day = ws2.cell(row=row, column=1)
            cell_day.alignment = align_center
            cell_day.border = border_thin
            cell_day.font = font_s2_day_lp
            for col in range(2, 19):
                c = ws2.cell(row=row, column=col)
                c.border = border_thin
                c.font = font_s2_cell_data
                c.alignment = align_center
            if day <= days_in_month:
                cell_day.value = day
                if QDate(year, month, day) in holidays_map:
                    for col in range(1, 19):
                        ws2.cell(row=row, column=col).fill = grey_fill
            else:
                cell_day.value = "X"
                for col in range(2, 19):
                    ws2.cell(row=row, column=col).value = "X"

        row_razem = 34
        ws2.row_dimensions[row_razem].height = row_height_s2
        cell_razem = ws2.cell(row=row_razem, column=1, value="razem")
        cell_razem.font = font_s2_razem
        cell_razem.alignment = align_center
        cell_razem.border = border_thin
        for col in range(2, 19):
            c = ws2.cell(row=row_razem, column=col)
            c.border = border_thin
            c.font = font_s2_razem

        footer_row = 35
        ws2.row_dimensions[footer_row].height = 77.4
        ws2.merge_cells(f'A{footer_row}:R{footer_row}')

        # --- PRZYWRÓCONA LEGENDA ---
        legend_text = (
            "Urlop wypoczynkowy, Opieka Art 188§1, Opieka zasiłek, L4, Urlop okolicznościowy, Urlop wypoczynkowy \"na żądanie\",  Urlop bezpłatny,  "
            "Dn - Wypłata nadgodzin, Wś - Wolne za pracę w święto, Wn - Wolne za nadgodziny, W5 - Wolne z tytułu 5-dniowego tygodnia pracy, "
            "W - Urlop wychowawczy, Um - Urlop macierzyński, ŚR - Świadczenia reabilitacyjne, Śs - Świadek w sądzie, Kr - Oddanie krwi., "
            "S-szkolenie z udzielania dziecku pierwszej pomocy, D-delegacje/podróże służbowe, P-dni wolne na poszukiwanie pracy, "
            "Urek-nieobecności za które zgodnie z par. 16 pkt 2 rozporządzenia w sprawie sposobu usprawiedliwienia nieobecności w pracy "
            "oraz udzielania pracownikom zwolnień od pracy, pracownicy mają prawo do uzyskania rekompensaty pieniężnej od właściwego organu,"
            "Sn - spóźnienie, N-nieusprawiedliwione nieobecności w pracy, Nun-nieobecność usprawiedliwiona niepłatna"
        )
        leg_cell = ws2[f'A{footer_row}']
        leg_cell.value = legend_text
        leg_cell.font = font_s2_legend
        leg_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        leg_cell.border = border_thin

        sign_row = footer_row + 1
        ws2.row_dimensions[sign_row].height = 13
        ws2.merge_cells(f'A{sign_row}:E{sign_row}')
        ws2[f'A{sign_row}'] = "PODPIS KADR"
        ws2[f'A{sign_row}'].alignment = align_center
        ws2[f'A{sign_row}'].font = font_s2_sign
        ws2[f'A{sign_row}'].border = None

        ws2.merge_cells(f'F{sign_row}:R{sign_row}')
        ws2[f'F{sign_row}'] = "PODPIS DYREKTORA ŻŁOBKA"
        ws2[f'F{sign_row}'].alignment = align_center
        ws2[f'F{sign_row}'].font = font_s2_sign
        ws2[f'F{sign_row}'].border = None

        frame_last_row = 35
        for col in range(1, 19):
            ws2.cell(row=1, column=col).border = Border(top=medium, bottom=thin, left=thin, right=thin)
        ws2['A35'].border = Border(top=thin, bottom=medium, left=medium, right=medium)
        for row in range(1, frame_last_row + 1):
            c = ws2.cell(row=row, column=1)
            curr = c.border
            c.border = Border(top=curr.top, bottom=curr.bottom, left=medium, right=curr.right)
        for row in range(1, frame_last_row + 1):
            c = ws2.cell(row=row, column=18)
            curr = c.border
            c.border = Border(top=curr.top, bottom=curr.bottom, left=curr.left, right=medium)

        filename = f"{emp['name'].replace(' ', '_')}_{month_str}_{year}.xlsx"
        path = os.path.join(folder, filename)
        wb.save(path)
        return path
